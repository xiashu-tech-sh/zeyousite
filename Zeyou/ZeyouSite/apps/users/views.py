from django.shortcuts import render
from random import randint
from django.core.mail import send_mail
from rest_framework.generics import CreateAPIView
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import JsonResponse, QueryDict
from django.forms.models import model_to_dict
from .models import User, email_verifica
from students.models import Student
from academies.models import Academy
from .serializers import CreateUserSerializer
from datetime import date, datetime
import json
from django.db.models import Q
import os
import openpyxl
import xlrd
import logging
loggcollect = logging.getLogger('file')


class UserView(CreateAPIView):
    """
    用户注册
    """
    serializer_class = CreateUserSerializer


class UpdatePwdView(APIView):

    def post(self, request):
        data = request.data
        user = User.objects.get(id=data['user_id'])
        if user is None:
            return Response({'message': "用户不存在"})
        if not user.check_password(data['old_password']):
            return Response({'message': "原密码输入有误"})
        if data['new_password'] != data['confirm_password']:
            return Response('两次密码不一致，请重新输入')
        user.set_password(data['new_password'])
        user.save()
        loggcollect.info("修改密码成功，用户名: %s"% user.username)
        return Response({'message': "修改成功", "status": 200})

class UpdatePerdView(APIView):

    def post(self, request):
        try:
            User.objects.filter(id=request.data["user_id"]).update(phone=request.data['phone'], email=request.data['email'])
            return Response({'message': "更新成功", "status": 200})
        except:
            return Response({'message': "邮箱已注册", "status": 202})


class DateEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.strftime("%Y-%m-%d")
        else:
            return json.JSONEncoder.default(self, obj)

class FilterDepartmentView(APIView):

    def post(self, request):
        data = request.data
        name = data['username']
        dict_infor = {}
        identity = data['identity']
        little_helper = data['little_helper']
        if identity:
            info = Student.objects.filter().all()
        elif little_helper:
            info = Student.objects.filter(Q(customer_state="未分配已购买") | Q(customer_state="未分配未购买")).all()
        else:
            department = User.objects.get(username=name).department
            if "助教" in department:
                update_shows = Academy.objects.filter(Q(warning_show='false')).all()
                for update_show in update_shows:
                    try:
                        infor_show = model_to_dict(update_show)
                        old_time = infor_show['follow_number']
                        new_time = datetime.now().strftime('%Y-%m-%d')
                        d1 = datetime.strptime(old_time, '%Y-%m-%d')
                        d2 = datetime.strptime(new_time, '%Y-%m-%d')
                        delta = d2 - d1
                        if infor_show['follow_period'] =='每周':
                            diff = 7
                        elif infor_show['follow_period'] =='每二周':
                            diff = 10
                        elif infor_show['follow_period'] == '每月':
                            diff = 30
                        if delta.days > diff:
                            Academy.objects.filter(order_number=infor_show['order_number']).update(warning_show='true')
                            Academy.objects.filter(name=infor_show['name']).update(warning_show='true')
                    except:
                        continue
                ty_data = Academy.objects.filter(Q(teaching_assistant=name)).values_list('name',flat = "true")
                list_ty_name = list(ty_data)
                if len(list_ty_name)==0:
                    info = Student.objects.filter(Q(little_assistant=name) | Q(consultant=name) |
                                                  Q(service_consultant=name) | Q(paper_writer=name)).all()

                else:
                    stu_name = Student.objects.filter(Q(little_assistant=name) | Q(consultant=name) |
                                                  Q(service_consultant=name) | Q(paper_writer=name)).values_list('name',flat = "true")
                    all_name = list(stu_name)
                    for ty_ in list_ty_name:
                        if ty_ not in all_name:
                            all_name.append(ty_)
                    info = Student.objects.filter(name=all_name[0]).all()
                    for _name in range(1,len(all_name)):
                        info = info | Student.objects.filter(name=all_name[_name]).all()

            else:
                info = Student.objects.filter(Q(little_assistant=name) | Q(consultant=name) |
                                          Q(service_consultant=name)| Q(paper_writer=name)).all()
        dict_infor['custpag'] = info.count()
        ret = []
        for i in range(info.count()):
            data1 = model_to_dict(info[i])
            data2 = dict()
            data2["id"] = data1["id"]
            data2["name"] = data1["name"]
            data2["graduation_date"] = data1["graduation_date"]
            data2["school"] = data1["school"]
            data2["warning_show"] = data1["warning_show"]
            ret.append(data2)
        dict_infor['data_infor'] = ret
        jsonArr = json.dumps(dict_infor, ensure_ascii=False)
        return JsonResponse(jsonArr, safe=False)

    def get(self,request,studentns_id):
        int_id = int(studentns_id)
        result = Student.objects.get(id=int_id)
        dist_result = model_to_dict(result)
        jsonArr = json.dumps(dist_result, cls = DateEncoder,ensure_ascii=False)
        return JsonResponse(jsonArr, safe=False)


class PersonalInfoView(APIView):

    def post(self, request):
        try:
            user = User.objects.filter(id=request.data["user_id"]).first()
            return Response({"name": user.username, "phone": user.phone, "email": user.email})
        except:
            return Response({"message": "user not found", "status": 400})


class LookupUser(APIView):

    def post(self, request):
        try:
            current_user = User.objects.filter(id=request.data["user_id"]).first()
            if current_user.department != 'A':
                return Response({"message": "Permission denied", "status": 400})
            user = User.objects.filter(request.data["name"]).first()
            user_dict = model_to_dict(user)
            return user_dict
        except:
            Response({"message": "User lookup error", "status": 400})


class DeleteUser(APIView):

    def post(self, request):
        try:
            current_user = User.objects.filter(id=request.data["user_id"]).first()
            if current_user.department != 'A':
                return Response({"message": "Permission denied", "status": 400})
            User.objects.filter(request.data["name"]).delete()
            return Response({"message": "User deletion successfully", "status": 201})
        except:
            Response({"message": "User lookup error", "status": 400})


class UserDetailView(APIView):

    # 更新用户信息
    def post(self, request):
        count = User.objects.filter(email=request.data["email"]).count()
        if count!= 0:
            return Response({'message': "当前邮箱已注册", 'status': 400})
        try:
            User.objects.filter(id=request.data["user_id"]).update(phone=request.data['phone'],email=request.data['email'],username=request.data['username'],)
            # 修改成功
            # Successfully modified
            return Response({'message': "用户信息更新成功",'status':200})
        except:
            # 参数不全
            # Incomplete parameters
            return Response({'message': "参数不完整",'status':400})

    # 获取用户信息
    def get(self, request, user_id):
        try:
            data = User.objects.get(id=user_id)
        except:
            # 非法请求
            return Response({'message': "非法请求",'status':201})
        data_dict = {
            "username": data.username,
            "email": data.email,
            "phone": data.phone,
            'status': 200

        }
        return Response(data_dict)

class ParseExcelView(APIView):
    def post(self, request):
        if os.path.exists('D:\\PythonProject\\Zeyou\\file\\dataset.xlsx') == False:
            return Response({'message': "No such file to be parsed", "status": 400})

        wb = openpyxl.load_workbook('D:\\PythonProject\\Zeyou\\file\\dataset.xlsx', data_only=True)
        allsheets = wb.sheetnames
        index = 0
        while index < 2:
            sheet = wb[allsheets[index]]
            maxrow = sheet.max_row

            if index == 0:
                #continue
                headers = ["customer_state", "source", "date_to_add", "name", "gender", "wechat_num", "area",
                           "phone", "little_assistant", "consultant", "service_consultant", "paper_writer", "identity",
                           "school", "school_type", "curriculum_system", "curriculum_system_note", "graduation_date",
                           "application_level", "major", "target_country", "GPA", "TOEFL", "IELTS", "SAT", "ACT", "GRE"]
                lists = []
                for row in range(2, maxrow + 1):
                    r = {}
                    for col in range(1, len(headers) + 1):
                        key = headers[col - 1]
                        r[key] = sheet.cell(row=row, column=col + 1).value
                        if key == 'gender':
                            if r[key] == '男':
                                r[key] = 'M'
                            else:
                                r[key] = 'F'

                        if key == 'identity':
                            if r[key] == '家长':
                                r[key] = 0
                            else:
                                r[key] = 1

                        if key == 'customer_state':
                            if r[key] == '未分配未购买':
                                r[key] = 0
                            if r[key] == '已分配未购买':
                                r[key] = 1
                            if r[key] == '未分配已购买':
                                r[key] = 2
                            if r[key] == '已分配已购买':
                                r[key] = 3
                            if r[key] == '已签约未购买':
                                r[key] = 4
                            if r[key] == '已签约已购买':
                                r[key] = 5

                        if key == 'date_to_add':
                            if r[key] is not None:
                                dtime = xlrd.xldate_as_tuple(r[key], 0)  # 转化为元组形式
                                d = date(dtime[0], dtime[1], dtime[2])
                                r[key] = d

                        if r[key] is None:
                            r[key] = ""
                    lists.append(r)

                sqllist = []
                for cell in lists:
                    # for header in headers:
                    customer_state = cell['customer_state']
                    source = cell['source']
                    date_to_add = cell['date_to_add']
                    name = cell['name']
                    gender = cell['gender']
                    wechat_num = cell['wechat_num']
                    area = cell['area']
                    phone = cell['phone']
                    little_assistant=cell["little_assistant"]
                    consultant = cell["consultant"]
                    service_consultant = cell["service_consultant"]
                    paper_writer = cell["paper_writer"]
                    identity = cell['identity']
                    school=cell["school"]
                    school_type=cell["school_type"]
                    curriculum_system = cell["curriculum_system"]
                    curriculum_system_note = cell["curriculum_system_note"]
                    application_level = cell["application_level"]
                    major = cell["major"]
                    target_country = cell["target_country"]
                    GPA = cell["GPA"]
                    TOEFL=cell["TOEFL"]
                    IELTS=cell["IELTS"]
                    SAT = cell["SAT"]
                    ACT = cell["ACT"]
                    GRE = cell["GRE"]
                    sql = Student(customer_state=customer_state, date_to_add=date_to_add, source=source, name=name,
                                  wechat_num=wechat_num, area=area, phone=phone, gender=gender, identity=identity,
                                  little_assistant=little_assistant, consultant=consultant, service_consultant=service_consultant,
                                  paper_writer=paper_writer, school=school, school_type=school_type,
                                  curriculum_system=curriculum_system, curriculum_system_note=curriculum_system_note,
                                  application_level=application_level, major=major, target_country=target_country,
                                  GPA=GPA, TOEFL=TOEFL, IELTS=IELTS, SAT=SAT, ACT=ACT, GRE=GRE)
                    sqllist.append(sql)
                try:
                    Student.objects.bulk_create(sqllist)
                except:
                    return Response({"message": "Parameters are incomplete and cannot be saved", "status": 400})
                index += 1

            elif index == 1:

                headers = ["name", "source", "product_type", "teaching_assistant", "sales", "teacher",
                           "date_of_purchasing", "product", "date_of_lecture", "hours_of_lecture", "price_per_hour"
                           , "price_overall", "cur_state"]
                lists = []
                for row in range(2, maxrow + 1):
                    r = {}
                    for col in range(1, len(headers) + 1):
                        key = headers[col - 1]
                        r[key] = sheet.cell(row=row, column=col).value

                        if key == 'hours_of_lecture' or key == 'price_per_hour' or key == 'price_overall':
                            if r[key] is not None:
                                r[key] = int(r[key])

                        if key == 'date_of_lecture' or key == 'date_of_purchasing':
                            if r[key] is not None:
                                dtime = xlrd.xldate_as_tuple(r[key], 0)  # 转化为元组形式
                                d = date(dtime[0], dtime[1], dtime[2])
                                r[key] = d
                        if r[key] is None and key != 'date_of_lecture' and key != 'date_of_purchasing':
                            r[key] = ""
                    lists.append(r)
                sqllist = []
                for cell in lists:
                    # for header in headers:
                    name = cell['name']
                    source = cell['source']
                    date_of_purchasing = cell['date_of_purchasing']
                    hours_of_lecture = cell['hours_of_lecture']
                    price_overall = cell["price_overall"]
                    date_of_lecture = cell['date_of_lecture']
                    product_type = cell['product_type']
                    product = cell['product']
                    price_per_hour = cell['price_per_hour']
                    cur_state = cell['cur_state']
                    teacher = cell['teacher']
                    teaching_assistant = cell['teaching_assistant']
                    sales = cell['sales']
                    sql = Academy(name=name, source=source, price_overall=price_overall, date_of_purchasing=date_of_purchasing,
                                  date_of_lecture=date_of_lecture, hours_of_lecture=hours_of_lecture, product=product,
                                  product_type=product_type, price_per_hour=price_per_hour, cur_state=cur_state, teacher=teacher,
                                  teaching_assistant=teaching_assistant, sales=sales)
                    sqllist.append(sql)
                try:
                    Academy.objects.bulk_create(sqllist)
                except:
                    return Response({"message": "Parameters are incomplete and cannot be saved", "status": 400})
                index += 1
        return Response({'message': "Successfully parsed", "status": 200})


class EmailCountView(APIView):
    """
    获取邮箱数量
    """
    def post(self, request):
        count = User.objects.filter(email=request.data["email"]).count()
        return Response({ 'count': count})


class UsernameCountView(APIView):
    """
    获取用户名数量
    """
    def post(self, request):
        count = User.objects.filter(username=request.data["username"]).count()
        return Response({ 'count': count})


class GetassistantView(APIView):
    def get(self, *args, **kwargs):
        ount = User.objects.filter().all()
        result = []
        for i in range(ount.count()):
            data1 = model_to_dict(ount[i])
            identity = data1['department'].split(',')
            for depar in identity:
                _dist = {}
                if depar =='小助手':
                    _dist['id'] = data1['id']
                    _dist['name'] = data1['username']
                    result.append(_dist)
        jsonArr = json.dumps(result, ensure_ascii=False)
        return JsonResponse(jsonArr, safe=False)


class ConsultantView(APIView):
    def get(self, *args, **kwargs):
        ount = User.objects.filter().all()
        result = []
        for i in range(ount.count()):
            data1 = model_to_dict(ount[i])
            identity = data1['department'].split(',')
            for depar in identity:
                _dist = {}
                if depar =='顾问':
                    _dist['id'] = data1['id']
                    _dist['name'] = data1['username']
                    result.append(_dist)
        jsonArr = json.dumps(result, ensure_ascii=False)
        return JsonResponse(jsonArr, safe=False)


class ServiceadvisorView(APIView):
    def get(self, *args, **kwargs):
        ount = User.objects.filter().all()
        result = []
        for i in range(ount.count()):
            data1 = model_to_dict(ount[i])
            identity = data1['department'].split(',')
            for depar in identity:
                _dist = {}
                if depar =='服务顾问':
                    _dist['id'] = data1['id']
                    _dist['name'] = data1['username']
                    result.append(_dist)
        jsonArr = json.dumps(result, ensure_ascii=False)
        return JsonResponse(jsonArr, safe=False)


class CopywritingView(APIView):
    def get(self, *args, **kwargs):
        ount = User.objects.filter().all()
        result = []
        for i in range(ount.count()):
            data1 = model_to_dict(ount[i])
            identity = data1['department'].split(',')
            for depar in identity:
                _dist = {}
                if depar =='文案':
                    _dist['id'] = data1['id']
                    _dist['name'] = data1['username']
                    result.append(_dist)
        jsonArr = json.dumps(result, ensure_ascii=False)
        return JsonResponse(jsonArr, safe=False)


class TeachingassistantView(APIView):

    def get(self, *args, **kwargs):
        ount = User.objects.filter().all()
        result = []
        for i in range(ount.count()):
            data1 = model_to_dict(ount[i])
            identity = data1['department'].split(',')
            for depar in identity:
                _dist = {}
                if depar =='助教':
                    _dist['id'] = data1['id']
                    _dist['name'] = data1['username']
                    result.append(_dist)
        jsonArr = json.dumps(result, ensure_ascii=False)
        return JsonResponse(jsonArr, safe=False)
class StrategyView(APIView):

    def get(self, *args, **kwargs):
        ount = User.objects.filter().all()
        result = []
        for i in range(ount.count()):
            data1 = model_to_dict(ount[i])
            identity = data1['department'].split(',')
            for depar in identity:
                _dist = {}
                if depar =='战略顾问':
                    _dist['id'] = data1['id']
                    _dist['name'] = data1['username']
                    result.append(_dist)
        jsonArr = json.dumps(result, ensure_ascii=False)
        return JsonResponse(jsonArr, safe=False)

class GetUserInfoView(APIView):

    def post(self,request):
        try:
            data = request.data
            User.objects.filter(id=data['id']).update(department=data['usertype'])
            return Response({'message': "更新成功", "status": 200})
        except:
            return Response({'message': "参数有误，更新失败", "status": 202})
    def get(self, *args, **kwargs):
        ount = User.objects.filter().all()
        result = []
        for i in range(ount.count()):
            user_dist = {}
            all_user_infor = model_to_dict(ount[i])
            user_dist['id'] = all_user_infor['id']
            user_dist['username'] = all_user_infor['username']
            user_dist['phone'] = all_user_infor['phone']
            user_dist['email'] = all_user_infor['email']
            user_dist['usertype'] = all_user_infor['department']
            result.append(user_dist)
        jsonArr = json.dumps(result, ensure_ascii=False)
        return JsonResponse(jsonArr, safe=False)

    def delete(self, request):
        data = request.data
        obj = User.objects.get(id=data["user_id"])
        obj.delete()
        # 删除成功
        return Response({"message": "删除成功","status":200})


class EmailCodeView(APIView):

    def post(self, request):
        data = email_verifica.objects.filter(email=request.data["email"])
        try:
            data.delete()
        except:
            pass
        try:
            user_infor = User.objects.get(email=request.data['email'])
        except:
            return Response({"message":"邮件地址不存在","status":202})

        email_code = '%06d' % randint(0, 999999)
        subject = "择由找回密码邮箱验证"
        html_message = '<p>尊敬的用户您好！</p>' \
                       '<p>感谢您使用择由智能客群管理系统。</p>' \
                       '<p>您的邮箱为：%s 。</p>' \
                       '<p>您的验证码为：%s</p>' % (request.data["email"],email_code)
        send_mail(subject, "", 'hhpbaby@126.com',[request.data["email"]], html_message=html_message)

        email_verifica.objects.create(email=request.data["email"],code=email_code)

        return Response({"message": "验证码已发送，请查收","status":200})


class EmailPasswordView(APIView):

    # 验证验证码
    def post(self, request):
        data = email_verifica.objects.filter(email=request.data["email"])
        if str(data.values()[0]['code']) != str(request.data['sms_code']):
            return Response({'message': "验证码错误","status":200})
        data.delete()
        if request.data['password'] != request.data['confirm_password']:
            return Response({'message': "两次密码输入不一致","status":200})
        user = User.objects.get(email=request.data["email"])
        user.set_password(request.data['password'])
        user.save()
        return Response({'message': "修改成功","status":200})
